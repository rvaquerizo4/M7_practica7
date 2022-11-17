import csv #importa el paquete csv

def escrbir_archivo(): #funcion
     #lista de nombre
    data = [
    ['Name', 'Surname' , 'Age'],
    ['Leonardo', 'Chavez', '34'],
    ['Raul', 'Vaquerizo', '45']
            ]

    with open('data.csv', 'w') as f: #define el nombre del archivo, indica que escribir y lo indica como una variable f
        csv_writer = csv.writer(f) #crea el archivo
        csv_writer.writerows(data) #writerows escribir simultaneo la lista


def lectura_archivo(): #llamada de funcion
    with open('data.csv', newline='') as f: #indica el nombre, y simbolo indica su signo de separacion
        reader = csv.reader(f) #abre el archivo
        for row in reader: #recorre el archivo linea por linea
            print(row) #imprimir

escrbir_archivo() #llamada de la funcion

print('-------------------------------')

lectura_archivo() #llamada de la funcion



print('--------------------------')

print('Creacion archivo excel')


from openpyxl import Workbook #importa los paquetes necesarios para excel
import openpyxl

def escritura():#funcion
    book = Workbook() # crear el archivo tipo excel
    sheet = book.active #abre el archivo

    #sheet ['A1'] indica la casilla donde se va a guardar
    sheet['A1'] = 'Name'
    sheet['A2'] = 'Herson'
    sheet['B1'] = 'Surname'
    sheet['B2'] = 'Vega'
    sheet['C1'] = 'Age'
    sheet['C2'] = '65'
    book.save("data.xlsx") #lo guarda en el archivo

def lectura ():
    book = openpyxl.load_workbook('data.xlsx') #carga el documento en la variable book

    sheet = book.active #importa los datos del archivo

    a1 = sheet['A1'] #guarda lo que existe en la casilla 1 en una variable
    a2 = sheet['A2']
    a3 = sheet.cell(row=3, column=1)

#imprimir datos
    print(a1.value)
    print(a2.value)
    print(a3.value)

escritura()

print('--------------------------')

lectura()